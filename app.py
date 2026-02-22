"""
╔═════════════════════════════════════════════════════════════════════════╗
║   QR Certificate System  v3.01                                           ║
║   Developed by: Abdul Samad | SBBU Nawabshah                            ║
╠═════════════════════════════════════════════════════════════════════════╣
║   pip install streamlit pillow qrcode[pil] reportlab openpyxl pandas    ║
║   streamlit run app.py                                                  ║
╚═════════════════════════════════════════════════════════════════════════╝
v7.0 NEW:
  ✅ Dual certificate source — Online Registrations OR Upload Names File
  ✅ File upload supports: Excel (.xlsx/.xls), CSV (.csv), TXT (.txt)
  ✅ Smart column detection — finds "name" column automatically
  ✅ Edit names before generating (manual corrections)
  ✅ Premium UI/UX — glassmorphism, smooth animations, responsive
  ✅ Better certificate designer with live settings panel
  ✅ Sample file downloads for guidance
"""

import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import qrcode
import io, zipfile, csv, os, json, base64, hashlib, hmac, secrets
import pandas as pd
import openpyxl
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.utils import ImageReader
from datetime import datetime

# ══════════════════════════════════════════════════════════════════
#  FILE PATHS
# ══════════════════════════════════════════════════════════════════
DATA_FILE   = "registrations.csv"
CONFIG_FILE = "config.json"
AUTH_FILE   = "auth.json"
BACKUP_DIR  = "backups"
CSV_HEADERS = ["ref_no","name","roll_no","department","batch",
               "category","event","date","time"]

# ══════════════════════════════════════════════════════════════════
#  SECURE PASSWORD  (PBKDF2-HMAC-SHA256)
# ══════════════════════════════════════════════════════════════════
def _hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(32)
    key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 310_000)
    return key.hex(), salt

def _verify_password(password, stored_hash, salt):
    candidate, _ = _hash_password(password, salt)
    return hmac.compare_digest(candidate, stored_hash)

def load_auth():
    if os.path.exists(AUTH_FILE):
        with open(AUTH_FILE) as f: return json.load(f)
    try:
        sec = st.secrets.get("auth", {})
        if sec.get("hash") and sec.get("salt"):
            return {"hash": sec["hash"], "salt": sec["salt"], "source": "secrets"}
    except: pass
    h, s = _hash_password("Admin@2025")
    auth = {"hash":h,"salt":s,"created":datetime.now().isoformat(),
            "note":"Change password immediately!"}
    try:
        with open(AUTH_FILE,"w") as f: json.dump(auth,f,indent=2)
    except: pass
    return auth

def save_password(new_password):
    h, s = _hash_password(new_password)
    auth = load_auth()
    auth.update({"hash":h,"salt":s,"changed":datetime.now().isoformat()})
    try:
        with open(AUTH_FILE,"w") as f: json.dump(auth,f,indent=2)
    except: pass
    st.session_state["_new_hash"] = h
    st.session_state["_new_salt"] = s

def check_password(password):
    auth = load_auth()
    return _verify_password(password, auth["hash"], auth["salt"])

# ══════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════
CFG_DEFAULTS = {
    "event_name":"Certificate of Participation","event_date":datetime.now().strftime("%Y-%m-%d"),
    "event_venue":"","event_topic":"","organizer":"",
    "categories":"Participant,Teacher,Speaker,Management",
    "student_cats":"Participant","app_url":"","inv_theme":"royal_gold",
    "logo1_b64":"","logo2_b64":"","logo3_b64":"",
}

def load_config():
    if not os.path.exists(CONFIG_FILE): return CFG_DEFAULTS.copy()
    try:
        with open(CONFIG_FILE,"r",encoding="utf-8") as f: saved=json.load(f)
        out=CFG_DEFAULTS.copy(); out.update(saved); return out
    except: return CFG_DEFAULTS.copy()

def save_config(cfg):
    with open(CONFIG_FILE,"w",encoding="utf-8") as f:
        json.dump(cfg,f,ensure_ascii=False,indent=2)

# ══════════════════════════════════════════════════════════════════
#  CSV DATABASE
# ══════════════════════════════════════════════════════════════════
def generate_ref_no(category):
    regs=load_registrations(); count=len(regs)+1
    words=category.strip().split()
    code=words[0][0].upper() if words else "R"
    if len(words)>1: code+=words[1][0].upper()
    return f"{code}-{count:04d}"

def save_registration(rec):
    exists=os.path.exists(DATA_FILE)
    with open(DATA_FILE,"a",newline="",encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=CSV_HEADERS)
        if not exists: w.writeheader()
        w.writerow({k:rec.get(k,"") for k in CSV_HEADERS})

def load_registrations():
    if not os.path.exists(DATA_FILE): return []
    try:
        with open(DATA_FILE,"r",encoding="utf-8") as f: return list(csv.DictReader(f))
    except: return []

def clear_registrations():
    if os.path.exists(DATA_FILE): os.remove(DATA_FILE)

# ══════════════════════════════════════════════════════════════════
#  BACKUP
# ══════════════════════════════════════════════════════════════════
def create_backup():
    buf=io.BytesIO(); ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    with zipfile.ZipFile(buf,"w",zipfile.ZIP_DEFLATED) as zf:
        for fname in [DATA_FILE,CONFIG_FILE]:
            if os.path.exists(fname): zf.write(fname,f"backup_{ts}/{fname}")
        zf.writestr(f"backup_{ts}/README.txt",
            f"Backup: {datetime.now().isoformat()}\nauth.json excluded for security.")
    return buf.getvalue()

def auto_backup():
    os.makedirs(BACKUP_DIR,exist_ok=True)
    bfile=os.path.join(BACKUP_DIR,f"backup_{datetime.now().strftime('%Y%m%d')}.zip")
    if not os.path.exists(bfile):
        with open(bfile,"wb") as f: f.write(create_backup())

# ══════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════
st.set_page_config(page_title="QR Certificate Generator Pro",
                   page_icon="🎓",layout="wide",
                   initial_sidebar_state="expanded")

# ══════════════════════════════════════════════════════════════════
#  PREMIUM CSS  — Glassmorphism + Responsive
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Base ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
.stApp{
  background:linear-gradient(160deg,#04091a 0%,#080f28 35%,#0d1535 70%,#111e45 100%);
  font-family:'Inter',sans-serif!important;min-height:100vh;}
/* ── Sidebar ── */
section[data-testid="stSidebar"]{
  background:linear-gradient(180deg,#060b20 0%,#0e1438 60%,#141c48 100%)!important;
  border-right:1px solid rgba(46,107,239,.2)!important;}
section[data-testid="stSidebar"] *{color:#a8c4f0!important;}
section[data-testid="stSidebar"] h2{
  color:#ffd159!important;font-size:1rem!important;
  border-bottom:1px solid #ffd15933;padding-bottom:8px;margin-bottom:12px;}
section[data-testid="stSidebar"] h3{color:#7ecefd!important;font-size:.9rem!important;}
/* ── Typography ── */
h1{color:#ffd159!important;text-align:center;font-weight:900!important;
   letter-spacing:-0.5px;text-shadow:0 0 40px rgba(255,209,89,.25);}
h2{color:#7ecefd!important;font-weight:800!important;}
h3{color:#a8d4ff!important;font-weight:700!important;}
h4{color:#7ecefd99!important;font-weight:600!important;}
p,li{color:#c0d4ee;}
/* ── Inputs ── */
.stTextInput>div>div>input,
.stNumberInput>div>div>input,
.stTextArea>div>div>textarea{
  background:rgba(8,20,48,.9)!important;color:#e8f2ff!important;
  border:1.5px solid rgba(46,107,239,.35)!important;
  border-radius:12px!important;padding:11px 16px!important;
  font-size:.93rem!important;transition:all .2s ease!important;}
.stTextInput>div>div>input:focus,
.stTextArea>div>div>textarea:focus{
  border-color:rgba(126,206,253,.7)!important;
  box-shadow:0 0 0 3px rgba(126,206,253,.12),0 4px 16px rgba(46,107,239,.15)!important;}
label,.stTextInput label,.stSelectbox label,.stSlider label,
.stTextArea label,.stCheckbox label,.stFileUploader label{
  color:#8bb8e0!important;font-weight:600!important;font-size:.87rem!important;
  letter-spacing:.3px!important;text-transform:uppercase!important;}
/* ── Selectbox ── */
.stSelectbox>div>div{
  background:rgba(8,20,48,.9)!important;color:#e8f2ff!important;
  border:1.5px solid rgba(46,107,239,.35)!important;border-radius:12px!important;}
/* ── Buttons ── */
.stButton>button{
  background:linear-gradient(135deg,#1a4aef 0%,#2d6ff5 50%,#5398f8 100%)!important;
  color:#fff!important;border:none!important;border-radius:12px!important;
  font-weight:700!important;font-size:.92rem!important;
  padding:.68rem 1.5rem!important;letter-spacing:.2px!important;
  box-shadow:0 4px 20px rgba(46,107,239,.3)!important;
  transition:all .25s cubic-bezier(.4,0,.2,1)!important;}
.stButton>button:hover{
  transform:translateY(-2px)!important;
  box-shadow:0 8px 28px rgba(46,107,239,.45)!important;
  filter:brightness(1.08)!important;}
.stButton>button:active{transform:translateY(0)!important;}
/* ── File uploader ── */
[data-testid="stFileUploader"]{
  background:rgba(8,20,48,.7)!important;
  border:2px dashed rgba(46,107,239,.4)!important;
  border-radius:16px!important;padding:8px!important;
  transition:border-color .2s!important;}
[data-testid="stFileUploader"]:hover{
  border-color:rgba(126,206,253,.6)!important;}
/* ── Cards ── */
.card{
  background:linear-gradient(135deg,rgba(12,20,52,.97),rgba(8,14,38,.98));
  border:1px solid rgba(46,107,239,.25);border-radius:20px;
  padding:28px;margin:10px 0;
  box-shadow:0 8px 32px rgba(0,0,0,.4),inset 0 1px 0 rgba(255,255,255,.05);}
.card-success{
  background:linear-gradient(135deg,rgba(6,40,24,.95),rgba(4,28,16,.98));
  border:1px solid rgba(46,204,113,.3);border-radius:16px;
  padding:18px;margin:8px 0;
  box-shadow:0 4px 20px rgba(46,204,113,.12);}
.card-warn{
  background:linear-gradient(135deg,rgba(55,28,0,.95),rgba(70,35,0,.98));
  border:1px solid rgba(243,156,18,.3);border-radius:16px;
  padding:18px;margin:8px 0;
  box-shadow:0 4px 20px rgba(243,156,18,.1);}
.card-info{
  background:linear-gradient(135deg,rgba(6,28,58,.95),rgba(4,18,42,.98));
  border:1px solid rgba(52,152,219,.3);border-radius:16px;
  padding:18px;margin:8px 0;
  box-shadow:0 4px 20px rgba(52,152,219,.1);}
.card-glass{
  background:rgba(255,255,255,.03);
  backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
  border:1px solid rgba(255,255,255,.07);border-radius:20px;
  padding:28px;margin:10px 0;
  box-shadow:0 8px 32px rgba(0,0,0,.3);}
/* ── Source toggle buttons ── */
.src-active{
  background:linear-gradient(135deg,#1a4aef,#5398f8)!important;
  border:2px solid rgba(126,206,253,.5)!important;
  box-shadow:0 6px 24px rgba(46,107,239,.4)!important;}
.src-inactive{
  background:rgba(46,107,239,.1)!important;
  border:2px solid rgba(46,107,239,.25)!important;
  box-shadow:none!important;}
/* ── Name chips ── */
.chip{
  display:inline-flex;align-items:center;gap:5px;
  background:rgba(46,107,239,.18);border:1px solid rgba(46,107,239,.4);
  border-radius:20px;padding:4px 12px;font-size:.82rem;
  color:#a8d4ff;margin:3px;}
.chip-more{
  background:rgba(255,209,89,.12);border-color:rgba(255,209,89,.3);
  color:#ffd159!important;}
/* ── Stat cards ── */
.stat{
  background:linear-gradient(135deg,rgba(12,20,52,.97),rgba(8,14,38,.98));
  border:1px solid rgba(46,107,239,.25);border-radius:16px;
  padding:20px 16px;text-align:center;
  box-shadow:0 4px 16px rgba(0,0,0,.3);}
.stat-n{font-size:2.4rem;font-weight:900;color:#ffd159;line-height:1.1;}
.stat-l{font-size:.8rem;color:#7ecefd;margin-top:4px;font-weight:600;
        text-transform:uppercase;letter-spacing:.5px;}
/* ── Metrics ── */
[data-testid="stMetricValue"]{
  color:#ffd159!important;font-size:2rem!important;font-weight:900!important;}
[data-testid="stMetricLabel"]{color:#7ecefd!important;font-weight:600!important;}
[data-testid="metric-container"]{
  background:linear-gradient(135deg,rgba(12,20,52,.97),rgba(8,14,38,.98))!important;
  border:1px solid rgba(46,107,239,.25)!important;
  border-radius:16px!important;padding:18px!important;}
/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"]{
  background:rgba(4,9,26,.9)!important;
  border-radius:16px 16px 0 0!important;
  padding:6px 6px 0!important;gap:3px!important;
  border-bottom:1px solid rgba(46,107,239,.2)!important;}
.stTabs [data-baseweb="tab"]{
  color:#7ecefd!important;background:transparent!important;
  border-radius:12px 12px 0 0!important;font-weight:700!important;
  font-size:.85rem!important;padding:10px 16px!important;
  transition:all .2s ease!important;border:none!important;}
.stTabs [data-baseweb="tab"]:hover{
  background:rgba(46,107,239,.2)!important;}
.stTabs [aria-selected="true"]{
  background:linear-gradient(135deg,#1a4aef,#2d6ff5)!important;
  color:#fff!important;
  box-shadow:0 4px 16px rgba(46,107,239,.4)!important;}
.stTabs [data-baseweb="tab-panel"]{
  background:rgba(4,9,26,.6)!important;
  border-radius:0 0 20px 20px!important;
  border:1px solid rgba(46,107,239,.15)!important;
  border-top:none!important;padding:20px!important;}
/* ── Progress ── */
.stProgress>div>div>div{
  background:linear-gradient(90deg,#1a4aef,#5398f8,#7ecefd)!important;
  border-radius:999px!important;transition:width .3s ease!important;}
/* ── Dataframe ── */
.stDataFrame{
  border-radius:14px!important;overflow:hidden!important;
  border:1px solid rgba(46,107,239,.25)!important;}
/* ── Expander ── */
.streamlit-expanderHeader{
  background:rgba(12,20,52,.8)!important;
  border-radius:12px!important;color:#7ecefd!important;
  font-weight:700!important;border:1px solid rgba(46,107,239,.2)!important;}
/* ── Checkboxes ── */
.stCheckbox>label{
  color:#a8c4f0!important;font-size:.9rem!important;font-weight:600!important;}
/* ── HR ── */
hr{border:none!important;
   border-top:1px solid rgba(46,107,239,.15)!important;margin:20px 0!important;}
/* ── Scrollbar ── */
::-webkit-scrollbar{width:5px;height:5px;}
::-webkit-scrollbar-track{background:rgba(4,9,26,.8);}
::-webkit-scrollbar-thumb{background:rgba(46,107,239,.5);border-radius:3px;}
::-webkit-scrollbar-thumb:hover{background:rgba(126,206,253,.6);}
/* ── Dev card styles ── */
.dev-card{
  background:linear-gradient(135deg,rgba(10,16,44,.98),rgba(6,10,30,.99));
  border:2px solid rgba(255,209,89,.4);border-radius:24px;
  padding:44px 36px;text-align:center;
  box-shadow:0 20px 60px rgba(0,0,0,.5),inset 0 1px 0 rgba(255,209,89,.1);}
.dev-name{font-size:2.8rem;font-weight:900;color:#ffd159;
          letter-spacing:2px;margin:14px 0 6px;
          text-shadow:0 0 30px rgba(255,209,89,.3);}
.dev-title{font-size:1.05rem;color:#7ecefd;letter-spacing:.8px;margin-bottom:8px;}
.soc-row{display:flex;gap:12px;justify-content:center;flex-wrap:wrap;margin:22px 0;}
.soc-btn{display:inline-flex;align-items:center;gap:8px;padding:11px 22px;
  border-radius:30px;font-weight:700;font-size:.9rem;
  text-decoration:none;transition:all .25s ease;}
.soc-btn:hover{transform:translateY(-3px);filter:brightness(1.12);}
.skill-row{display:flex;gap:9px;flex-wrap:wrap;justify-content:center;margin:14px 0;}
.skill{background:rgba(46,107,239,.25);border:1px solid rgba(46,107,239,.5);
  color:#a8d4ff;padding:6px 16px;border-radius:20px;
  font-size:.87rem;font-weight:600;}
.dev-hr{border:none!important;border-top:1px solid rgba(255,209,89,.15)!important;margin:20px 0!important;}
/* ── Responsive ── */
@media(max-width:768px){
  .stTabs [data-baseweb="tab"]{font-size:.75rem!important;padding:8px 8px!important;}
  .card,.card-glass{padding:16px!important;}
  h1{font-size:1.6rem!important;}
  .dev-name{font-size:2rem!important;}
  .soc-row{gap:8px!important;}
  .soc-btn{padding:9px 14px!important;font-size:.82rem!important;}
}
@media(max-width:480px){
  .stTabs [data-baseweb="tab"]{font-size:.7rem!important;padding:7px 6px!important;}
  h1{font-size:1.3rem!important;}
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
#  STARTUP
# ══════════════════════════════════════════════════════════════════
_cfg = load_config()
auto_backup()

# ══════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════
SESS = {
    "admin_auth":False,"template_bytes":None,"qr_data":None,
    "event_name":_cfg["event_name"],"event_date":_cfg["event_date"],
    "event_venue":_cfg["event_venue"],"event_topic":_cfg["event_topic"],
    "organizer":_cfg["organizer"],"categories":_cfg["categories"],
    "student_cats_input":_cfg["student_cats"],"app_url":_cfg["app_url"],
    "inv_theme":_cfg["inv_theme"],
    "logo1_b64":_cfg["logo1_b64"],"logo2_b64":_cfg["logo2_b64"],"logo3_b64":_cfg["logo3_b64"],
    "text_x":50,"text_y":60,"font_size":72,"text_color":"#1a1a1a","selected_font":"Arial Bold",
    "form_submitted":False,"last_submission":{},"invitation_png":None,
    "cert_source":"registered","uploaded_names":[],
}
for k,v in SESS.items():
    if k not in st.session_state: st.session_state[k]=v

# ══════════════════════════════════════════════════════════════════
#  FONTS
# ══════════════════════════════════════════════════════════════════
FONTS = {
    "Arial Bold":["arialbd.ttf","DejaVuSans-Bold.ttf"],
    "Arial Regular":["arial.ttf","DejaVuSans.ttf"],
    "Arial Italic":["ariali.ttf","DejaVuSans-Oblique.ttf"],
    "Calibri Bold":["calibrib.ttf","DejaVuSans-Bold.ttf"],
    "Calibri Regular":["calibri.ttf","DejaVuSans.ttf"],
    "Tahoma Bold":["tahomabd.ttf","DejaVuSans-Bold.ttf"],
    "Tahoma Regular":["tahoma.ttf","DejaVuSans.ttf"],
    "Verdana Bold":["verdanab.ttf","DejaVuSans-Bold.ttf"],
    "Verdana Regular":["verdana.ttf","DejaVuSans.ttf"],
    "Trebuchet Bold":["trebucbd.ttf","DejaVuSans-Bold.ttf"],
    "Trebuchet MS":["trebuc.ttf","DejaVuSans.ttf"],
    "Segoe UI Bold":["segoeuib.ttf","DejaVuSans-Bold.ttf"],
    "Segoe UI":["segoeui.ttf","DejaVuSans.ttf"],
    "Times New Roman Bold":["timesbd.ttf","DejaVuSerif-Bold.ttf"],
    "Times New Roman":["times.ttf","DejaVuSerif.ttf"],
    "Times New Roman Italic":["timesi.ttf","DejaVuSerif-Italic.ttf"],
    "Georgia Bold":["georgiab.ttf","DejaVuSerif-Bold.ttf"],
    "Georgia Regular":["georgia.ttf","DejaVuSerif.ttf"],
    "Palatino Bold":["palab.ttf","DejaVuSerif-Bold.ttf"],
    "Palatino Linotype":["pala.ttf","DejaVuSerif.ttf"],
    "Garamond Bold":["GARABD.TTF","DejaVuSerif-Bold.ttf"],
    "Garamond":["GARA.TTF","DejaVuSerif.ttf"],
    "Courier New Bold":["courbd.ttf","DejaVuSansMono-Bold.ttf"],
    "Courier New":["cour.ttf","DejaVuSansMono.ttf"],
    "Consolas Bold":["consolab.ttf","DejaVuSansMono-Bold.ttf"],
    "Consolas":["consola.ttf","DejaVuSansMono.ttf"],
    "Century Gothic Bold":["GOTHICB.TTF","DejaVuSans-Bold.ttf"],
    "Century Gothic":["GOTHIC.TTF","DejaVuSans.ttf"],
    "Impact":["impact.ttf","DejaVuSans-Bold.ttf"],
    "Rockwell Bold":["ROCKB.TTF","DejaVuSerif-Bold.ttf"],
    "Rockwell":["ROCK.TTF","DejaVuSerif.ttf"],
    "Brush Script MT":["BRUSHSCI.TTF","DejaVuSerif-Italic.ttf"],
    "Lucida Handwriting":["lhandw.ttf","DejaVuSerif-Italic.ttf"],
    "Comic Sans Bold":["comicbd.ttf","DejaVuSans-Bold.ttf"],
    "DejaVu Sans Bold":["DejaVuSans-Bold.ttf","DejaVuSans-Bold.ttf"],
    "DejaVu Sans":["DejaVuSans.ttf","DejaVuSans.ttf"],
    "DejaVu Serif Bold":["DejaVuSerif-Bold.ttf","DejaVuSerif-Bold.ttf"],
    "DejaVu Serif":["DejaVuSerif.ttf","DejaVuSerif.ttf"],
    "DejaVu Mono Bold":["DejaVuSansMono-Bold.ttf","DejaVuSansMono-Bold.ttf"],
}
FONT_CATS = {
    "🔤 Sans Serif":[k for k in FONTS if any(x in k for x in ["Arial","Calibri","Tahoma","Verdana","Trebuchet","Segoe"])],
    "📜 Serif / Formal":[k for k in FONTS if any(x in k for x in ["Times","Georgia","Palatino","Garamond"])],
    "💻 Monospace":[k for k in FONTS if any(x in k for x in ["Courier","Consolas"])],
    "✨ Display":[k for k in FONTS if any(x in k for x in ["Century","Impact","Rockwell"])],
    "🖋️ Script":[k for k in FONTS if any(x in k for x in ["Brush","Handwriting","Comic"])],
    "🛡️ DejaVu":[k for k in FONTS if "DejaVu" in k],
}

# ══════════════════════════════════════════════════════════════════
#  INVITATION CARD THEMES
# ══════════════════════════════════════════════════════════════════
THEMES = {
    "royal_gold":{"bg":(12,8,32),"bg2":(28,18,60),"bg3":(45,30,90),
        "acc":(255,215,0),"acc2":(255,180,50),"txt":(255,255,255),
        "sub":(220,190,255),"brd":(140,90,220),"bbg":(255,215,0),"btxt":(20,10,50),"label":"✨ Royal Gold"},
    "midnight_blue":{"bg":(5,15,40),"bg2":(10,30,70),"bg3":(15,50,100),
        "acc":(100,180,255),"acc2":(60,140,220),"txt":(255,255,255),
        "sub":(160,210,255),"brd":(46,107,239),"bbg":(46,107,239),"btxt":(255,255,255),"label":"🌙 Midnight Blue"},
    "crimson_elite":{"bg":(25,5,10),"bg2":(55,10,20),"bg3":(80,20,30),
        "acc":(255,80,100),"acc2":(220,50,70),"txt":(255,255,255),
        "sub":(255,180,190),"brd":(180,40,60),"bbg":(200,30,50),"btxt":(255,255,255),"label":"🔴 Crimson Elite"},
    "emerald_prestige":{"bg":(5,22,15),"bg2":(8,45,28),"bg3":(12,70,42),
        "acc":(50,220,130),"acc2":(30,180,100),"txt":(255,255,255),
        "sub":(150,240,200),"brd":(30,160,90),"bbg":(20,150,80),"btxt":(255,255,255),"label":"💚 Emerald Prestige"},
    "obsidian_gold":{"bg":(8,8,8),"bg2":(20,18,14),"bg3":(30,26,18),
        "acc":(212,175,55),"acc2":(180,140,30),"txt":(255,255,255),
        "sub":(200,185,140),"brd":(180,145,40),"bbg":(180,145,40),"btxt":(10,8,5),"label":"⚫ Obsidian Gold"},
    "ocean_sapphire":{"bg":(3,25,40),"bg2":(5,50,70),"bg3":(8,75,100),
        "acc":(0,210,200),"acc2":(0,170,165),"txt":(255,255,255),
        "sub":(120,235,230),"brd":(0,160,155),"bbg":(0,180,175),"btxt":(5,30,40),"label":"🌊 Ocean Sapphire"},
    "violet_luxury":{"bg":(18,5,40),"bg2":(35,10,75),"bg3":(55,20,110),
        "acc":(200,130,255),"acc2":(170,100,230),"txt":(255,255,255),
        "sub":(220,180,255),"brd":(130,60,210),"bbg":(130,60,210),"btxt":(255,255,255),"label":"💜 Violet Luxury"},
    "rose_gold":{"bg":(30,12,18),"bg2":(55,22,30),"bg3":(80,35,45),
        "acc":(240,170,120),"acc2":(210,140,95),"txt":(255,255,255),
        "sub":(255,210,190),"brd":(190,100,70),"bbg":(200,110,75),"btxt":(255,255,255),"label":"🌹 Rose Gold"},
}
THEME_LABELS = {k:v["label"] for k,v in THEMES.items()}

def get_invite_phrase(category):
    c=category.lower()
    if any(x in c for x in ["teacher","professor","faculty","lecturer","principal"]):
        return "You are cordially invited as"
    elif any(x in c for x in ["speaker","keynote","presenter"]):
        return "We are honored to welcome"
    elif any(x in c for x in ["chief","director","ceo","vip","guest of honor"]):
        return "It is our privilege to invite"
    elif any(x in c for x in ["judge","panelist","reviewer","evaluator"]):
        return "You are invited to serve as"
    elif any(x in c for x in ["business","entrepreneur","sponsor","investor"]):
        return "We are pleased to welcome"
    elif any(x in c for x in ["management","organizer","volunteer","coordinator"]):
        return "You are invited to participate as"
    elif any(x in c for x in ["alumni","graduate"]):
        return "We warmly welcome our distinguished alumnus"
    return "We are pleased to invite"

# ══════════════════════════════════════════════════════════════════
#  PIL HELPERS
# ══════════════════════════════════════════════════════════════════
def _fnt(size, bold=False):
    cands=(["arialbd.ttf","DejaVuSans-Bold.ttf","calibrib.ttf"]
           if bold else ["arial.ttf","DejaVuSans.ttf","calibri.ttf"])
    for f in cands:
        try: return ImageFont.truetype(f,size)
        except: pass
    return ImageFont.load_default()

def _rr(draw,x1,y1,x2,y2,r,fill,outline=None,ow=2):
    if x2<=x1 or y2<=y1: return
    r=min(r,(x2-x1)//2,(y2-y1)//2)
    draw.rectangle([x1+r,y1,x2-r,y2],fill=fill)
    draw.rectangle([x1,y1+r,x2,y2-r],fill=fill)
    for ex,ey in [(x1,y1),(x2-2*r,y1),(x1,y2-2*r),(x2-2*r,y2-2*r)]:
        draw.ellipse([ex,ey,ex+2*r,ey+2*r],fill=fill)
    if outline:
        for ex,ey,s,e in [(x1,y1,180,270),(x2-2*r,y1,270,360),
                           (x1,y2-2*r,90,180),(x2-2*r,y2-2*r,0,90)]:
            draw.arc([ex,ey,ex+2*r,ey+2*r],s,e,fill=outline,width=ow)
        draw.line([x1+r,y1,x2-r,y1],fill=outline,width=ow)
        draw.line([x1+r,y2,x2-r,y2],fill=outline,width=ow)
        draw.line([x1,y1+r,x1,y2-r],fill=outline,width=ow)
        draw.line([x2,y1+r,x2,y2-r],fill=outline,width=ow)

def _grad(draw,x1,y1,x2,y2,c1,c2,vert=True):
    steps=(y2-y1) if vert else (x2-x1)
    for i in range(max(1,steps)):
        a=i/max(1,steps-1)
        col=tuple(int(c1[j]*(1-a)+c2[j]*a) for j in range(3))
        if vert: draw.line([(x1,y1+i),(x2,y1+i)],fill=col)
        else: draw.line([(x1+i,y1),(x1+i,y2)],fill=col)

def _wrap(draw,text,font,max_w):
    words=text.split(); lines=[]; cur=""
    for w in words:
        test=(cur+" "+w).strip()
        if draw.textbbox((0,0),test,font=font)[2]>max_w:
            if cur: lines.append(cur)
            cur=w
        else: cur=test
    if cur: lines.append(cur)
    return lines or [text]

def load_pil_font(name,size):
    for path in FONTS.get(name,["DejaVuSans-Bold.ttf"]):
        try: return ImageFont.truetype(path,size)
        except: pass
    return ImageFont.load_default()

def hex_rgba(h,a=255):
    h=h.lstrip("#")
    return (int(h[0:2],16),int(h[2:4],16),int(h[4:6],16),a)

# ══════════════════════════════════════════════════════════════════
#  INVITATION CARD GENERATOR
# ══════════════════════════════════════════════════════════════════
def generate_invitation_card(rec,cfg,l1=None,l2=None,l3=None):
    W,H=1080,1620
    th=THEMES.get(cfg.get("inv_theme","royal_gold"),THEMES["royal_gold"])
    bg=th["bg"];bg2=th["bg2"];bg3=th["bg3"]
    acc=th["acc"];acc2=th["acc2"];txt=th["txt"];sub=th["sub"];brd=th["brd"]
    bbg=th["bbg"];btxt=th["btxt"]
    img=Image.new("RGB",(W,H),bg); draw=ImageDraw.Draw(img)
    mid=H//2
    _grad(draw,0,0,W,mid,bg,bg2); _grad(draw,0,mid,W,H,bg2,bg3)
    for cx,cy,cr in [(150,200,280),(W-120,H-300,320),(W//2,H//2,400)]:
        for dr in range(0,cr,28):
            draw.ellipse([cx-dr,cy-dr,cx+dr,cy+dr],outline=(*brd[:3],10),width=1)
    for xi in range(-200,W+200,110):
        draw.line([(xi,0),(xi+280,H)],fill=(*brd[:3],12),width=1)
    _rr(draw,16,16,W-16,H-16,36,bg2,outline=acc,ow=3)
    _rr(draw,26,26,W-26,H-26,28,bg,outline=brd,ow=1)
    _grad(draw,16,16,W-16,30,acc2,acc,vert=False)
    _grad(draw,16,H-30,W-16,H-16,acc,acc2,vert=False)
    for cx,cy in [(16+28,16+28),(W-16-28,16+28),(16+28,H-16-28),(W-16-28,H-16-28)]:
        sz=18
        draw.polygon([(cx,cy-sz),(cx+sz,cy),(cx,cy+sz),(cx-sz,cy)],fill=acc)
        draw.polygon([(cx,cy-sz//2),(cx+sz//2,cy),(cx,cy+sz//2),(cx-sz//2,cy)],fill=bg2)
    y=70
    LH=105; lraw=[b for b in [l1,l2,l3] if b]; limgs=[]
    for lb in lraw:
        try:
            li=Image.open(io.BytesIO(lb)).convert("RGBA")
            r=LH/li.height; li=li.resize((max(1,int(li.width*r)),LH),Image.LANCZOS)
            limgs.append(li)
        except: pass
    if limgs:
        gap=44; total=sum(l.width for l in limgs)+(len(limgs)-1)*gap
        xs=(W-total)//2
        for li in limgs: img.paste(li,(xs,y),li); xs+=li.width+gap
    else:
        draw.text((W//2,y+LH//2),"🎓",font=_fnt(88),fill=acc,anchor="mm")
    y+=LH+16
    org=cfg.get("organizer","")
    if org:
        draw.text((W//2,y),org.upper(),font=_fnt(26,True),fill=acc,anchor="mt"); y+=38
    _grad(draw,70,y,W-70,y+3,acc2,acc,vert=False); y+=14
    draw.text((W//2,y),"✦  I N V I T A T I O N  ✦",font=_fnt(23),fill=sub,anchor="mt"); y+=46
    ev_font=_fnt(55,True); ev_lines=_wrap(draw,cfg.get("event_name","Event"),ev_font,W-140)
    for ln in ev_lines:
        draw.text((W//2,y),ln,font=ev_font,fill=acc,anchor="mt"); y+=66
    y+=4
    topic=cfg.get("event_topic","")
    if topic:
        tp_font=_fnt(26); tb=draw.textbbox((0,0),topic,font=tp_font)
        tw_=tb[2]-tb[0]+64; _rr(draw,W//2-tw_//2,y,W//2+tw_//2,y+46,23,brd)
        draw.text((W//2,y+23),topic,font=tp_font,fill=sub,anchor="mm"); y+=60
    y+=8
    draw.rectangle([90,y,W-90,y+1],fill=(*brd[:3],100)); y+=22
    phrase=get_invite_phrase(rec.get("category",""))
    draw.text((W//2,y),phrase,font=_fnt(27),fill=sub,anchor="mt"); y+=44
    name=rec.get("name","Participant"); nm_font=_fnt(60,True)
    nm_lines=_wrap(draw,name,nm_font,W-120); name_h=len(nm_lines)*74+32
    _rr(draw,44,y,W-44,y+name_h,22,bg3,outline=acc,ow=3)
    ny=y+(name_h-len(nm_lines)*74)//2
    for ln in nm_lines:
        draw.text((W//2,ny+37),ln,font=nm_font,fill=acc,anchor="mm"); ny+=74
    y+=name_h+14
    cat_text=rec.get("category",""); cb=draw.textbbox((0,0),cat_text,font=_fnt(30,True))
    cw=cb[2]-cb[0]+90; bx1=W//2-cw//2; bx2=W//2+cw//2
    _grad(draw,bx1,y,bx2,y+56,acc,acc2,vert=False)
    draw.text((W//2,y+28),cat_text,font=_fnt(30,True),fill=btxt,anchor="mm"); y+=70
    y+=10; dets=[]
    if rec.get("department"): dets.append(("🏛  Department",rec["department"]))
    if rec.get("roll_no"):    dets.append(("🔢  Roll No",rec["roll_no"]))
    if rec.get("batch"):      dets.append(("📅  Batch",rec["batch"]))
    if dets:
        box_h=len(dets)*56+24
        _rr(draw,46,y,W-46,y+box_h,18,bg2,outline=brd,ow=1)
        dy=y+20
        for i,(lbl,val) in enumerate(dets):
            draw.text((100,dy),lbl,font=_fnt(24),fill=sub,anchor="lt")
            draw.text((W-100,dy),val,font=_fnt(26,True),fill=txt,anchor="rt")
            if i<len(dets)-1:
                draw.line([(100,dy+38),(W-100,dy+38)],fill=(*brd[:3],70),width=1)
            dy+=56
        y+=box_h+16
    y+=6; evd=cfg.get("event_date","")
    try: evd=datetime.strptime(evd,"%Y-%m-%d").strftime("%B %d, %Y  (%A)")
    except: pass
    ev_items=[(i,v) for i,v in [("📅  Date",evd),("📍  Venue",cfg.get("event_venue","")),
                                  ("🏛  Organizer",cfg.get("organizer",""))] if v]
    if ev_items:
        ev_h=len(ev_items)*50+24; _rr(draw,46,y,W-46,y+ev_h,18,bg,outline=acc,ow=2)
        ey=y+18
        for icon,val in ev_items:
            elines=_wrap(draw,f"{icon}:  {val}",_fnt(24),W-140)
            for ln in elines: draw.text((W//2,ey),ln,font=_fnt(24),fill=sub,anchor="mt"); ey+=34
        y+=ev_h+16
    y+=6; ref=rec.get("ref_no","—"); reg_text=f"Reg No:  {ref}"
    rf_font=_fnt(32,True); rb=draw.textbbox((0,0),reg_text,font=rf_font)
    rw=rb[2]-rb[0]+80; rx1=W//2-rw//2; rx2=W//2+rw//2
    _rr(draw,rx1,y,rx2,y+62,31,bbg)
    _grad(draw,rx1+4,y+4,rx2-4,y+18,(*txt,55),(*txt,0))
    draw.text((W//2,y+31),reg_text,font=rf_font,fill=btxt,anchor="mm"); y+=78
    draw.text((W//2,y),"✦  Officially Registered  ✦",font=_fnt(22),fill=sub,anchor="mt")
    bar_y=H-62
    _grad(draw,16,bar_y,W-16,H-16,brd,bg2)
    _grad(draw,16,bar_y,W-16,bar_y+3,acc,acc2,vert=False)
    draw.text((W//2,bar_y+(H-16-bar_y)//2),
        f"{cfg.get('organizer','')}  •  {cfg.get('event_date','')}",
        font=_fnt(21),fill=acc,anchor="mm")
    buf=io.BytesIO(); img.save(buf,format="PNG",dpi=(150,150))
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
#  CERTIFICATE GENERATOR
# ══════════════════════════════════════════════════════════════════
def generate_cert(name,template,cfg_c):
    img=Image.open(io.BytesIO(template)).convert("RGBA")
    w,h=img.size; font=load_pil_font(cfg_c["font"],cfg_c["size"])
    px=int(w*cfg_c["x"]/100); py=int(h*cfg_c["y"]/100)
    layer=Image.new("RGBA",img.size,(255,255,255,0))
    draw=ImageDraw.Draw(layer)
    bbox=draw.textbbox((0,0),name,font=font)
    tw,th2=bbox[2]-bbox[0],bbox[3]-bbox[1]
    draw.text((px-tw//2,py-th2//2),name,font=font,fill=hex_rgba(cfg_c["color"]))
    out=Image.alpha_composite(img,layer).convert("RGB")
    buf=io.BytesIO(); out.save(buf,format="PNG",dpi=(300,300))
    return buf.getvalue()

def cert_to_pdf(png,name):
    buf=io.BytesIO(); pw,ph=landscape(A4)
    c=pdf_canvas.Canvas(buf,pagesize=(pw,ph))
    img=Image.open(io.BytesIO(png)).convert("RGB")
    iw,ih=img.size; sc=min(pw/iw,ph/ih); nw,nh=iw*sc,ih*sc
    tmp=io.BytesIO(); img.save(tmp,format="PNG"); tmp.seek(0)
    c.drawImage(ImageReader(tmp),(pw-nw)/2,(ph-nh)/2,nw,nh,mask="auto")
    c.setFont("Helvetica-Bold",9); c.setFillColorRGB(.5,.5,.5)
    c.drawCentredString(pw/2,14,
        f"{name}  |  {st.session_state.event_name}  |  {datetime.now().strftime('%Y-%m-%d')}")
    c.save(); return buf.getvalue()

def make_qr(url):
    qr=qrcode.QRCode(version=1,error_correction=qrcode.constants.ERROR_CORRECT_H,
                     box_size=10,border=4)
    qr.add_data(url); qr.make(fit=True)
    buf=io.BytesIO()
    qr.make_image(fill_color="#0b132b",back_color="white").save(buf,format="PNG")
    return buf.getvalue()

def cur_cfg():
    return {"x":st.session_state.text_x,"y":st.session_state.text_y,
            "size":st.session_state.font_size,"color":st.session_state.text_color,
            "font":st.session_state.selected_font}

# ══════════════════════════════════════════════════════════════════
#  NAMES FILE PARSER
# ══════════════════════════════════════════════════════════════════
def parse_names_file(uploaded_file):
    """Parse Excel, CSV, or TXT file and return list of names."""
    names = []
    ext = uploaded_file.name.rsplit(".",1)[-1].lower()
    try:
        if ext in ["xlsx","xls"]:
            df = pd.read_excel(uploaded_file, header=None)
            raw = df.iloc[:,0].dropna().astype(str).tolist()
            # Skip header if present
            if raw and raw[0].strip().lower() in ["name","full name","names","naam",
                                                   "student name","participant","sno","sr"]:
                raw = raw[1:]
            names = [n.strip() for n in raw if n.strip() and len(n.strip())>1]

        elif ext == "csv":
            df = pd.read_csv(uploaded_file)
            name_col = None
            for col in df.columns:
                if col.strip().lower() in ["name","full name","names","naam",
                                           "student name","participant","fullname"]:
                    name_col = col; break
            if name_col:
                names = df[name_col].dropna().astype(str).str.strip().tolist()
            else:
                # First column fallback
                names = df.iloc[:,0].dropna().astype(str).str.strip().tolist()
                # Remove first row if it looks like a header
                if names and names[0].lower() in ["name","full name","names"]:
                    names = names[1:]

        elif ext == "txt":
            content = uploaded_file.read().decode("utf-8","ignore")
            names = [l.strip() for l in content.splitlines() if l.strip()]

        # Final clean — remove obvious non-names
        names = [n for n in names
                 if n and len(n)>1
                 and n.lower() not in ["name","full name","names","naam","sr","sno","no","#"]]
    except Exception as e:
        return [], str(e)
    return names, None

# ══════════════════════════════════════════════════════════════════
#  EXCEL REPORT
# ══════════════════════════════════════════════════════════════════
def build_excel(regs):
    wb=openpyxl.Workbook()
    hf=PatternFill("solid",fgColor="1E1B4B"); hf2=PatternFill("solid",fgColor="0B132B")
    hfn=XFont(bold=True,color="FFFFFF",size=12)
    bdr=Border(bottom=Side(style="thin",color="334466"))
    ws=wb.active; ws.title="Registrations"
    ws.merge_cells("A1:I1"); t=ws["A1"]
    t.value=f"  {st.session_state.event_name} — Registration Data"
    t.font=XFont(bold=True,color="FFD159",size=14); t.fill=hf2
    t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=34
    ws.merge_cells("A2:I2"); info=ws["A2"]
    try: day=datetime.strptime(st.session_state.event_date,"%Y-%m-%d").strftime("%A")
    except: day=""
    info.value=(f"Date:{st.session_state.event_date}({day}) | "
                f"Venue:{st.session_state.event_venue} | "
                f"Organizer:{st.session_state.organizer} | Total:{len(regs)}")
    info.font=XFont(color="7ECEFD",size=10); info.fill=hf
    info.alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=18
    cols=[("Ref No",12),("#",5),("Full Name",28),("Roll No",14),
          ("Department",22),("Batch",12),("Category",16),("Date",14),("Time",10)]
    for ci,(h,w) in enumerate(cols,1):
        cell=ws.cell(row=3,column=ci,value=h)
        cell.font=hfn; cell.fill=hf
        cell.alignment=Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=22
    for ri,rec in enumerate(regs,4):
        alt=PatternFill("solid",fgColor="0F1B35" if ri%2==0 else "1A2550")
        vals=[rec.get("ref_no",""),ri-3,rec.get("name",""),rec.get("roll_no",""),
              rec.get("department",""),rec.get("batch",""),rec.get("category",""),
              rec.get("date",""),rec.get("time","")]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.font=XFont(color="E0E0E0",size=11); c.fill=alt; c.border=bdr
            c.alignment=Alignment(horizontal="center" if ci in[1,2,7,8,9] else "left",vertical="center")
        ws.row_dimensions[ri].height=20
    ws2=wb.create_sheet("Summary"); ws2.merge_cells("A1:C1"); t2=ws2["A1"]
    t2.value="Category Summary"; t2.font=XFont(bold=True,color="FFD159",size=13)
    t2.fill=hf2; t2.alignment=Alignment(horizontal="center"); ws2.row_dimensions[1].height=28
    for ci,h in enumerate(["Category","Count","Members"],1):
        c=ws2.cell(row=2,column=ci,value=h); c.font=hfn; c.fill=hf
        c.alignment=Alignment(horizontal="center")
    cats={}
    for rec in regs: cats.setdefault(rec.get("category","Other"),[]).append(f"{rec.get('name','')}")
    for ri,(cat,names) in enumerate(cats.items(),3):
        ws2.cell(row=ri,column=1,value=cat).font=XFont(bold=True,color="FFD159")
        ws2.cell(row=ri,column=2,value=len(names)).font=XFont(color="E0E0E0")
        ws2.cell(row=ri,column=3,value=", ".join(names)).font=XFont(color="E0E0E0")
        for col in range(1,4): ws2.cell(row=ri,column=col).fill=hf
    ws2.column_dimensions["A"].width=20; ws2.column_dimensions["B"].width=10
    ws2.column_dimensions["C"].width=80
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def save_all_settings():
    save_config({
        "event_name":st.session_state.event_name,"event_date":st.session_state.event_date,
        "event_venue":st.session_state.event_venue,"event_topic":st.session_state.event_topic,
        "organizer":st.session_state.organizer,"categories":st.session_state.categories,
        "student_cats":st.session_state.student_cats_input,"app_url":st.session_state.app_url,
        "inv_theme":st.session_state.inv_theme,
        "logo1_b64":st.session_state.logo1_b64,
        "logo2_b64":st.session_state.logo2_b64,
        "logo3_b64":st.session_state.logo3_b64,
    })

# ══════════════════════════════════════════════════════════════════
#  ROUTING
# ══════════════════════════════════════════════════════════════════
page = st.query_params.get("page","admin")

# ══════════════════════════════════════════════════════════════════
#  STUDENT FORM PAGE
# ══════════════════════════════════════════════════════════════════
if page == "form":
    cfg=load_config(); event=cfg.get("event_name","Certificate Event")
    cats=[c.strip() for c in cfg.get("categories","Participant").split(",") if c.strip()]
    s_cats=[c.strip().lower() for c in cfg.get("student_cats","Participant").split(",")]
    l1b=base64.b64decode(cfg["logo1_b64"]) if cfg.get("logo1_b64") else None
    l2b=base64.b64decode(cfg["logo2_b64"]) if cfg.get("logo2_b64") else None
    l3b=base64.b64decode(cfg["logo3_b64"]) if cfg.get("logo3_b64") else None

    st.markdown(f"""
    <div style="text-align:center;padding:28px 0 10px;">
      <div style="font-size:3.2rem;margin-bottom:6px;">🎓</div>
      <h1 style="color:#ffd159;font-size:2rem;margin:6px 0;text-shadow:0 0 30px rgba(255,209,89,.3);">
        {event}
      </h1>
      <p style="color:#7ecefd;margin:4px 0;font-size:.95rem;">
        {'📍 '+cfg.get('event_venue','') if cfg.get('event_venue') else ''}
        {'&nbsp;|&nbsp; 📅 '+cfg.get('event_date','') if cfg.get('event_date') else ''}
      </p>
      <p style="color:#7ecefd66;font-size:.85rem;">
        Organized by {cfg.get('organizer','')}
      </p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    # ── CONFIRMATION SCREEN ───────────────────────────────────────
    if st.session_state.get("form_submitted") and st.session_state.get("invitation_png"):
        rec=st.session_state.last_submission; inv_png=st.session_state.invitation_png

        st.markdown("""
        <div style="text-align:center;padding:8px 0 16px;">
          <div style="font-size:3rem;">🎉</div>
          <h2 style="color:#2ecc71;margin:6px 0;">Registration Successful!</h2>
          <p style="color:#7ecefd;font-size:1rem;">
            Your Invitation Card is ready — download &amp; share as image!
          </p>
        </div>
        """, unsafe_allow_html=True)

        _,mc,_=st.columns([1,3,1])
        with mc: st.image(inv_png,use_container_width=True)

        st.markdown("---")
        ref=rec.get("ref_no","")
        fn=f"Invitation_{rec.get('name','').replace(' ','_')}_{ref}.png"
        d1,d2=st.columns(2)
        with d1:
            st.download_button("⬇️  Download Invitation Card",
                data=inv_png,file_name=fn,mime="image/png",use_container_width=True)
        with d2:
            wa=(f"🎓 *{event}*%0A%0A"
                f"Successfully registered!%0A"
                f"👤 *Name:* {rec.get('name','')}%0A"
                f"🏷️ *Category:* {rec.get('category','')}%0A"
                f"🆔 *Reg No:* {ref}%0A"
                f"📅 *Date:* {cfg.get('event_date','')}%0A"
                f"📍 *Venue:* {cfg.get('event_venue','')}%0A%0A"
                f"_(Download invitation card image above and share!)_")
            st.markdown(
                f'<a href="https://api.whatsapp.com/send?text={wa}" target="_blank"'
                f' style="display:block;text-align:center;padding:.7rem;border-radius:12px;'
                f'background:linear-gradient(90deg,#25D366,#128C7E);color:white;'
                f'font-weight:700;text-decoration:none;font-size:.95rem;">'
                f'📲 Share on WhatsApp</a>',unsafe_allow_html=True)

        st.markdown("""
        <div class="card-info" style="text-align:center;padding:14px;margin-top:8px;">
          💡 <b>To share as image:</b> Download → Open WhatsApp/Instagram/Facebook → Attach image → Send 📸
        </div>
        """,unsafe_allow_html=True)

        st.markdown("#### 🔗 Social Media")
        s1,s2,s3=st.columns(3)
        app_url=cfg.get("app_url","")
        with s1:
            st.markdown(f'<a href="https://www.facebook.com/sharer/sharer.php?u={app_url}" target="_blank" style="display:block;text-align:center;padding:.6rem;border-radius:10px;background:#1877F2;color:white;font-weight:700;text-decoration:none;">📘 Facebook</a>',unsafe_allow_html=True)
        with s2:
            st.markdown(f'<a href="https://www.linkedin.com/sharing/share-offsite/?url={app_url}" target="_blank" style="display:block;text-align:center;padding:.6rem;border-radius:10px;background:#0A66C2;color:white;font-weight:700;text-decoration:none;">💼 LinkedIn</a>',unsafe_allow_html=True)
        with s3:
            tw=f"I just registered for {event}! Reg No: {ref}"
            st.markdown(f'<a href="https://twitter.com/intent/tweet?text={tw}" target="_blank" style="display:block;text-align:center;padding:.6rem;border-radius:10px;background:#1DA1F2;color:white;font-weight:700;text-decoration:none;">🐦 Twitter</a>',unsafe_allow_html=True)

        st.markdown("---")
        with st.expander("📋 Registration Details"):
            r=rec
            st.markdown(f"""
| Field | Value |
|-------|-------|
| 🆔 Reg No | `{r.get('ref_no','')}` |
| 👤 Name | {r.get('name','')} |
| 🏷️ Category | {r.get('category','')} |
| 🏫 Department | {r.get('department','—')} |
| 🔢 Roll No | {r.get('roll_no','—')} |
| 📅 Batch | {r.get('batch','—')} |
| 🗓️ Date | {r.get('date','')} |
| 🕐 Time | {r.get('time','')} |
""")
        if st.button("🔄 New Registration",use_container_width=True):
            st.session_state.form_submitted=False
            st.session_state.last_submission={}
            st.session_state.invitation_png=None
            st.rerun()

    # ── FORM ─────────────────────────────────────────────────────
    else:
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("#### 📝 Fill Your Details")
        c1,c2=st.columns(2)
        with c1:
            name=st.text_input("👤 Full Name ✱",placeholder="Muhammad Ali Khan")
            dept=st.text_input("🏫 Department / Organization",
                                placeholder="Computer Science / ABC Company")
        with c2:
            category=st.selectbox("🏷️ Category ✱",cats)
            is_stud=category.lower() in s_cats
            rollno=st.text_input(
                "🔢 Roll No"+(" ✱" if is_stud else " (Optional)"),
                placeholder="CS-2022-45" if is_stud else "Optional")
            batch=st.text_input(
                "📅 Batch / Year"+(" ✱" if is_stud else " (Optional)"),
                placeholder="2022-2026") if is_stud else ""
        st.markdown("---")
        if st.button("✅  Submit Registration",use_container_width=True):
            n=name.strip(); r=rollno.strip(); d=dept.strip()
            b=batch.strip() if batch else ""
            missing=[]
            if not n: missing.append("Full Name")
            if is_stud and not r: missing.append("Roll No")
            if is_stud and not b: missing.append("Batch")
            if missing:
                st.error("❌ Required: **"+"  |  ".join(missing)+"**")
            else:
                now=datetime.now(); ref_no=generate_ref_no(category)
                rec={"ref_no":ref_no,"name":n,"roll_no":r,"department":d,
                     "batch":b,"category":category,"event":event,
                     "date":now.strftime("%Y-%m-%d"),"time":now.strftime("%H:%M:%S")}
                save_registration(rec)
                inv_png=generate_invitation_card(rec,cfg,l1b,l2b,l3b)
                st.session_state.form_submitted=True
                st.session_state.last_submission=rec
                st.session_state.invitation_png=inv_png
                st.rerun()
        st.markdown('</div>',unsafe_allow_html=True)

    st.markdown('<p style="text-align:center;color:#7ecefd22;font-size:.78rem;margin-top:24px;">Developed by Abdul Samad — SBBU Nawabshah</p>',unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════════════
#  ADMIN PAGE
# ══════════════════════════════════════════════════════════════════
st.markdown("# 🎓 QR Certificate Generator Pro")
st.markdown('<p style="text-align:center;color:#7ecefd88;margin-top:-8px;">v7.0 — Abdul Samad | SBBU Nawabshah</p>',unsafe_allow_html=True)
st.markdown("---")

if not st.session_state.admin_auth:
    _,col,_=st.columns([1,1.5,1])
    with col:
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("### 🔐 Admin Login")
        if not os.path.exists(AUTH_FILE):
            st.markdown('<div class="card-warn">🔑 <b>First Run!</b> Default password: <code>Admin@2025</code> — Change after login!</div>',unsafe_allow_html=True)
        pwd=st.text_input("Password",type="password",placeholder="Enter admin password")
        if st.button("🔓 Login",use_container_width=True):
            if check_password(pwd):
                st.session_state.admin_auth=True; st.rerun()
            else:
                st.error("❌ Incorrect password!")
        st.markdown('</div>',unsafe_allow_html=True)
    st.stop()

# ── Sidebar ─────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📋 Event Settings")
    st.session_state.event_name  =st.text_input("Event Name",         st.session_state.event_name)
    st.session_state.event_topic =st.text_input("Topic",              st.session_state.event_topic)
    st.session_state.event_date  =st.text_input("Date (YYYY-MM-DD)",  st.session_state.event_date)
    st.session_state.event_venue =st.text_input("Venue",              st.session_state.event_venue)
    st.session_state.organizer   =st.text_input("Organizer",          st.session_state.organizer)
    st.session_state.categories  =st.text_input("Categories (comma)", st.session_state.categories)
    st.session_state.student_cats_input=st.text_input("Student categories (Roll No required)",st.session_state.student_cats_input)
    st.markdown("---")
    st.markdown("## 🌐 App URL")
    st.session_state.app_url=st.text_input("Deployed URL",value=st.session_state.app_url,placeholder="https://yourapp.streamlit.app")
    if st.button("💾 Save All Settings",use_container_width=True):
        save_all_settings(); st.success("✅ Saved!")
    st.markdown("---")
    st.markdown("## 🖊️ Certificate Text")
    st.session_state.font_size =st.slider("Font Size",20,250,st.session_state.font_size)
    st.session_state.text_x   =st.slider("Horizontal % (←→)",0,100,st.session_state.text_x)
    st.session_state.text_y   =st.slider("Vertical %   (↑↓)",0,100,st.session_state.text_y)
    st.session_state.text_color=st.color_picker("Text Color",st.session_state.text_color)
    st.markdown("---")
    st.markdown("## 🎨 Invitation Theme")
    st.session_state.inv_theme=st.selectbox("Theme",list(THEMES.keys()),
        format_func=lambda x:THEME_LABELS[x],
        index=list(THEMES.keys()).index(st.session_state.inv_theme if st.session_state.inv_theme in THEMES else "royal_gold"))
    st.markdown("**Logos (up to 3):**")
    for li,lkey in enumerate(["logo1_b64","logo2_b64","logo3_b64"],1):
        lupl=st.file_uploader(f"Logo {li}",type=["png","jpg","jpeg"],key=f"lu{li}")
        if lupl:
            st.session_state[lkey]=base64.b64encode(lupl.read()).decode()
            st.success(f"✅ Logo {li}!")
        elif st.session_state.get(lkey):
            try: st.image(base64.b64decode(st.session_state[lkey]),width=65)
            except: pass
            if st.button(f"🗑️ Remove Logo {li}",key=f"rm{li}"):
                st.session_state[lkey]=""; st.rerun()
    st.markdown("---")
    st.markdown("## 🔤 Font")
    sq=st.text_input("🔍 Search...",placeholder="bold, times, gothic")
    if sq.strip():
        matched=[f for f in FONTS if sq.strip().lower() in f.lower()]
        if matched:
            idx=matched.index(st.session_state.selected_font) if st.session_state.selected_font in matched else 0
            st.session_state.selected_font=st.selectbox("Results:",matched,index=idx,key="fss")
        else: st.warning("No match")
    else:
        for cl,cf in FONT_CATS.items():
            if not cf: continue
            with st.expander(cl,expanded="Sans" in cl):
                for fn in cf:
                    lbl=("✅ " if st.session_state.selected_font==fn else "")+fn
                    if st.button(lbl,key=f"fb_{fn}",use_container_width=True):
                        st.session_state.selected_font=fn; st.rerun()
    st.markdown(f"**Selected:** `{st.session_state.selected_font}`")
    st.markdown("---")
    with st.expander("🔑 Change Password"):
        st.caption("8+ chars, uppercase, numbers, symbols")
        cur_p=st.text_input("Current",type="password",key="cp")
        new_p=st.text_input("New",type="password",key="np")
        cnf_p=st.text_input("Confirm",type="password",key="cfp")
        if st.button("🔒 Update Password",use_container_width=True):
            if not check_password(cur_p): st.error("❌ Wrong current password!")
            elif len(new_p)<8: st.error("❌ Min 8 characters!")
            elif new_p!=cnf_p: st.error("❌ Passwords don't match!")
            elif new_p==cur_p: st.warning("⚠️ Same as current!")
            else:
                save_password(new_p); st.success("✅ Updated!")
                nh=st.session_state.get("_new_hash",""); ns=st.session_state.get("_new_salt","")
                if nh:
                    st.markdown("**☁️ Streamlit Cloud → Settings → Secrets:**")
                    st.code(f'[auth]\nhash = "{nh}"\nsalt = "{ns}"',language="toml")
    if st.button("🚪 Logout"):
        st.session_state.admin_auth=False; st.rerun()

# ── Tabs ────────────────────────────────────────────────────────
tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8,tab9=st.tabs([
    "🔳 QR Generate","📊 Registrations","🃏 Invitation Card",
    "🖼️ Certificate Designer","🚀 Bulk Generate",
    "💾 Backup & Security","☁️ Deploy Guide","👨‍💻 Developer","📖 README",
])

# ─── TAB 1 — QR ─────────────────────────────────────────────────
with tab1:
    c1,c2=st.columns(2)
    with c1:
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("### 🔳 Registration QR Code")
        if st.session_state.app_url:
            st.markdown(f'<div class="card-info">✅ <b>URL:</b><br><code style="color:#ffd159;font-size:.85rem;">{st.session_state.app_url}</code></div>',unsafe_allow_html=True)
            qr_url=f"{st.session_state.app_url.rstrip('/')}/?page=form"
            if st.button("🔳 Refresh QR"): st.session_state.qr_data=make_qr(qr_url)
            if not st.session_state.qr_data: st.session_state.qr_data=make_qr(qr_url)
            _,qc,_=st.columns([1,2,1])
            with qc: st.image(st.session_state.qr_data,use_container_width=True)
            st.download_button("⬇️ Download QR PNG",st.session_state.qr_data,
                file_name="registration_qr.png",mime="image/png",use_container_width=True)
            st.code(qr_url,language=None)
        else:
            st.markdown('<div class="card-warn">⚠️ Set App URL in sidebar → Save Settings</div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("### 📱 How It Works")
        st.markdown("""
| Step | Action |
|------|--------|
| 1 | 📱 Scan QR code with phone |
| 2 | 📝 Fill Name, Dept, Roll No |
| 3 | 🏷️ Select Category |
| 4 | ✅ Submit |
| 5 | 🎉 Invitation Card instantly! |
| 6 | 📲 Download & Share as image |

**Works for:** Students • Teachers • Speakers  
Businessmen • Guests • VIPs • Alumni...
        """)
        st.markdown('</div>',unsafe_allow_html=True)
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("### ✏️ Manual Entry")
        with st.form("mf"):
            m1,m2=st.columns(2)
            with m1: mn=st.text_input("Name"); md=st.text_input("Department")
            with m2: mr=st.text_input("Roll No"); mb=st.text_input("Batch")
            mcl=[c.strip() for c in st.session_state.categories.split(",") if c.strip()]
            mc=st.selectbox("Category",mcl)
            if st.form_submit_button("➕ Add Manually",use_container_width=True):
                if mn.strip():
                    now=datetime.now(); ref=generate_ref_no(mc)
                    save_registration({"ref_no":ref,"name":mn.strip(),"roll_no":mr.strip(),
                        "department":md.strip(),"batch":mb.strip(),"category":mc,
                        "event":st.session_state.event_name,
                        "date":now.strftime("%Y-%m-%d"),"time":now.strftime("%H:%M:%S")})
                    st.success(f"✅ {mn} added!")
                else: st.error("Name required!")
        st.markdown('</div>',unsafe_allow_html=True)

# ─── TAB 2 — Registrations ──────────────────────────────────────
with tab2:
    regs=load_registrations()
    st.markdown("### 📊 Registration Data")
    col_r,col_b=st.columns([3,1])
    with col_r:
        cat_list=[c.strip() for c in st.session_state.categories.split(",") if c.strip()]
        mcols=st.columns(min(len(cat_list)+1,6))
        mcols[0].metric("Total",len(regs))
        for i,cat in enumerate(cat_list[:5]):
            mcols[i+1].metric(cat,sum(1 for r in regs if r.get("category","")==cat))
    with col_b:
        if st.button("🔄 Refresh",use_container_width=True): st.rerun()
    st.markdown("---")
    if regs:
        df=pd.DataFrame(regs)
        rename={"ref_no":"Reg No","name":"Full Name","roll_no":"Roll No",
                "department":"Department","batch":"Batch","category":"Category",
                "event":"Event","date":"Date","time":"Time"}
        df=df.rename(columns={k:v for k,v in rename.items() if k in df.columns})
        fc=st.selectbox("Filter:",["All"]+cat_list,key="flt")
        st.dataframe(df if fc=="All" else df[df["Category"]==fc],
                     use_container_width=True,height=380)
        st.markdown("---")
        e1,e2,e3=st.columns(3)
        with e1:
            st.download_button("📊 Export Excel",build_excel(regs),
                file_name=f"{st.session_state.event_name.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with e2:
            st.download_button("📄 Export TXT",
                "\n".join(f"{r.get('ref_no','')} | {r['name']} | {r.get('roll_no','')} | {r.get('category','')}" for r in regs).encode(),
                file_name="registrations.txt",mime="text/plain",use_container_width=True)
        with e3:
            if st.button("🗑️ Clear All",use_container_width=True):
                clear_registrations(); st.success("Cleared!"); st.rerun()
    else:
        st.info("📭 No registrations yet.")

# ─── TAB 3 — Invitation Card ────────────────────────────────────
with tab3:
    st.markdown("### 🃏 Invitation Card — Preview & Batch")
    cfg_n=load_config(); cfg_n["inv_theme"]=st.session_state.inv_theme
    l1b=base64.b64decode(st.session_state.logo1_b64) if st.session_state.logo1_b64 else None
    l2b=base64.b64decode(st.session_state.logo2_b64) if st.session_state.logo2_b64 else None
    l3b=base64.b64decode(st.session_state.logo3_b64) if st.session_state.logo3_b64 else None
    st.markdown(f"""
    <div class="card-info" style="padding:12px 18px;">
      🎨 Theme: <b style="color:#ffd159;">{THEME_LABELS.get(st.session_state.inv_theme,'—')}</b>
      &nbsp;|&nbsp; Logos: <b style="color:#ffd159;">{sum(1 for x in [l1b,l2b,l3b] if x)}</b> uploaded
      &nbsp;|&nbsp; <small>Change in sidebar → Save Settings</small>
    </div>
    """,unsafe_allow_html=True)
    pc1,pc2,pc3=st.columns(3)
    with pc1: pname=st.text_input("Preview name:","Muhammad Ali Khan",key="inv_pn")
    with pc2:
        pcatl=[c.strip() for c in st.session_state.categories.split(",") if c.strip()]
        pcat=st.selectbox("Category:",pcatl,key="inv_pc")
    with pc3: proll=st.text_input("Roll No:","CS-2022-45",key="inv_pr")
    pdept=st.text_input("Department:","Computer Science",key="inv_pd")
    srec={"ref_no":"P-0001","name":pname,"roll_no":proll,"department":pdept,
          "batch":"2022-2026","category":pcat,"event":st.session_state.event_name,
          "date":datetime.now().strftime("%Y-%m-%d")}
    iprev=generate_invitation_card(srec,cfg_n,l1b,l2b,l3b)
    _,mid,_=st.columns([1,3,1])
    with mid: st.image(iprev,use_container_width=True)
    pd1,pd2=st.columns(2)
    with pd1:
        st.download_button("⬇️ Download Preview",iprev,
            file_name=f"Preview_{pname.replace(' ','_')}.png",
            mime="image/png",use_container_width=True)
    with pd2:
        if st.button("💾 Save Theme Settings",use_container_width=True):
            save_all_settings(); st.success("✅ Saved!")
    st.markdown("---")
    regs_inv=load_registrations()
    if regs_inv:
        if st.button(f"🚀 Generate All {len(regs_inv)} Invitation Cards (ZIP)",use_container_width=True):
            p=st.progress(0); s=st.empty(); bz=io.BytesIO()
            with zipfile.ZipFile(bz,"w",zipfile.ZIP_DEFLATED) as zf:
                for i,rec in enumerate(regs_inv):
                    s.markdown(f"⏳ **{rec.get('name','')}** ({i+1}/{len(regs_inv)})")
                    card=generate_invitation_card(rec,cfg_n,l1b,l2b,l3b)
                    zf.writestr(f"Invitations/{rec.get('category','Other')}/{rec.get('ref_no','')}-{rec.get('name','')}.png",card)
                    p.progress((i+1)/len(regs_inv))
            s.success("✅ Done!")
            st.download_button("⬇️ All Invitation Cards ZIP",bz.getvalue(),
                file_name="All_Invitations.zip",mime="application/zip",use_container_width=True)
    else: st.info("No registrations yet.")

# ─── TAB 4 — Certificate Designer ───────────────────────────────
with tab4:
    st.markdown("### 🖼️ Certificate Designer")

    # ── Upload template ──────────────────────────────────────────
    st.markdown('<div class="card">',unsafe_allow_html=True)
    st.markdown("#### 📤 Certificate Template")
    st.caption("Upload your designed certificate background. Names will be placed automatically.")
    tc1,tc2=st.columns([3,2])
    with tc1:
        tpl=st.file_uploader("",type=["png","jpg","jpeg"],key="tpl_up",
                             label_visibility="collapsed")
        if tpl:
            st.session_state.template_bytes=tpl.read()
            img_t=Image.open(io.BytesIO(st.session_state.template_bytes))
            st.markdown(f'<div class="card-success">✅ <b>{tpl.name}</b> — {img_t.width}×{img_t.height}px</div>',unsafe_allow_html=True)
        if st.session_state.template_bytes:
            st.image(st.session_state.template_bytes,use_container_width=True)
    with tc2:
        if st.session_state.template_bytes:
            st.markdown("**⚙️ Current Text Settings**")
            settings=[("🔤 Font",st.session_state.selected_font[:22]),
                      ("📏 Size",f"{st.session_state.font_size}px"),
                      ("↔️ X Position",f"{st.session_state.text_x}%"),
                      ("↕️ Y Position",f"{st.session_state.text_y}%"),
                      ("🎨 Color",st.session_state.text_color)]
            for lbl,val in settings:
                st.markdown(f"""
                <div style="display:flex;justify-content:space-between;align-items:center;
                  padding:9px 12px;margin:4px 0;background:rgba(46,107,239,.1);
                  border-radius:8px;border:1px solid rgba(46,107,239,.2);">
                  <span style="color:#7ecefd;font-size:.88rem;">{lbl}</span>
                  <span style="color:#ffd159;font-weight:700;font-size:.88rem;">{val}</span>
                </div>""",unsafe_allow_html=True)
            st.markdown('<br><small style="color:#7ecefd88;">Adjust from sidebar sliders</small>',unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="text-align:center;padding:40px 20px;color:#7ecefd66;">
              <div style="font-size:3rem;margin-bottom:12px;">📋</div>
              <b>Upload template on the left</b><br>
              <small>PNG or JPG accepted</small>
            </div>""",unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)

    # ── Live preview ─────────────────────────────────────────────
    if st.session_state.template_bytes:
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("#### 👁️ Live Preview")
        pv_c1,pv_c2=st.columns([2,1])
        with pv_c1:
            pn=st.text_input("Preview name:","Muhammad Ali Khan",key="cpn")
        with pv_c2:
            st.markdown("<br>",unsafe_allow_html=True)
        pp=generate_cert(pn,st.session_state.template_bytes,cur_cfg())
        _,pc,_=st.columns([1,4,1])
        with pc: st.image(pp,use_container_width=True)
        dl1,dl2=st.columns(2)
        with dl1: st.download_button("⬇️ PNG",pp,file_name=f"Sample_{pn}.png",mime="image/png",use_container_width=True)
        with dl2: st.download_button("⬇️ PDF",cert_to_pdf(pp,pn),file_name=f"Sample_{pn}.pdf",mime="application/pdf",use_container_width=True)
        st.markdown('</div>',unsafe_allow_html=True)

        # ── Preview registered names ──────────────────────────
        regs_p=load_registrations()
        if regs_p:
            st.markdown('<div class="card">',unsafe_allow_html=True)
            st.markdown("#### 👥 Preview Registered Names")
            names_all=[r["name"] for r in regs_p]
            sn=st.slider("How many to preview?",1,min(len(names_all),24),min(6,len(names_all)))
            for i in range(0,sn,3):
                rn=names_all[i:i+3]; cs=st.columns(3)
                for ci,nm in enumerate(rn):
                    with cs[ci]:
                        pv=generate_cert(nm,st.session_state.template_bytes,cur_cfg())
                        st.image(pv,caption=nm,use_container_width=True)
                        st.download_button(f"⬇️ {nm[:16]}",pv,
                            file_name=f"{nm}.png",mime="image/png",
                            key=f"pv_{i}_{ci}",use_container_width=True)
            st.markdown('</div>',unsafe_allow_html=True)

# ─── TAB 5 — Bulk Generate ──────────────────────────────────────
with tab5:
    st.markdown("### 🚀 Bulk Certificate Generation")

    # ── Template check ────────────────────────────────────────────
    if not st.session_state.template_bytes:
        st.markdown('<div class="card-warn"><b>⚠️ No template uploaded!</b> Go to <b>🖼️ Certificate Designer</b> tab first.</div>',unsafe_allow_html=True)
        st.stop()

    # ── Source toggle ─────────────────────────────────────────────
    st.markdown("""
    <div style="margin-bottom:16px;">
      <p style="color:#7ecefd;font-weight:700;font-size:1rem;margin:0 0 12px;">
        📋 Select Name Source
      </p>
    </div>""",unsafe_allow_html=True)

    src_col1,src_col2=st.columns(2)
    with src_col1:
        is_reg=(st.session_state.cert_source=="registered")
        if st.button(
            "🌐  Online Registrations" + ("  ✅" if is_reg else ""),
            key="btn_src_reg",use_container_width=True,
            type="primary" if is_reg else "secondary"):
            st.session_state.cert_source="registered"; st.rerun()
    with src_col2:
        is_upl=(st.session_state.cert_source=="uploaded")
        if st.button(
            "📁  Upload Names File" + ("  ✅" if is_upl else ""),
            key="btn_src_upl",use_container_width=True,
            type="primary" if is_upl else "secondary"):
            st.session_state.cert_source="uploaded"; st.rerun()

    # ── Source description ────────────────────────────────────────
    if st.session_state.cert_source=="registered":
        st.markdown('<div class="card-info" style="padding:10px 16px;margin:8px 0 16px;"><small>📌 Using names from QR registrations (online form submissions)</small></div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="card-info" style="padding:10px 16px;margin:8px 0 16px;"><small>📌 Upload Excel/CSV/TXT file with names — for existing data</small></div>',unsafe_allow_html=True)

    st.markdown("---")

    # ════════════════════════════════════════════════════
    #  SOURCE A — ONLINE REGISTRATIONS
    # ════════════════════════════════════════════════════
    if st.session_state.cert_source=="registered":
        regs=load_registrations()
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("#### 🌐 Certificates from Online Registrations")

        if not regs:
            st.markdown('<div class="card-warn">⚠️ No registrations yet! Share QR code or use Manual Entry.</div>',unsafe_allow_html=True)
        else:
            cat_list_r=list(set(r.get("category","") for r in regs))
            sc=st.columns(min(len(cat_list_r)+1,5))
            sc[0].metric("Total",len(regs))
            for i,cat in enumerate(cat_list_r[:4]):
                sc[i+1].metric(cat[:12],sum(1 for r in regs if r.get("category","")==cat))
            st.markdown("---")
            fc_opts=["All Categories"]+cat_list_r
            fc=st.selectbox("Filter by category:",fc_opts,key="bulk_flt")
            filtered=[r for r in regs if fc=="All Categories" or r.get("category","")==fc]

            st.markdown(f"""
            <div class="card-info" style="padding:12px;text-align:center;margin:10px 0;">
              <b style="color:#ffd159;font-size:1.2rem;">{len(filtered)}</b>
              <span style="color:#7ecefd;"> certificates will be generated</span>
            </div>""",unsafe_allow_html=True)

            f1,f2=st.columns(2)
            with f1: do_png=st.checkbox("✅ PNG (Image)",value=True,key="dp1")
            with f2: do_pdf=st.checkbox("✅ PDF (Print quality)",value=True,key="dp2")

            if st.button(f"🚀 Generate {len(filtered)} Certificates",use_container_width=True):
                cn=cur_cfg(); prog=st.progress(0); stat=st.empty(); bz=io.BytesIO()
                with zipfile.ZipFile(bz,"w",zipfile.ZIP_DEFLATED) as zf:
                    for i,rec in enumerate(filtered):
                        nm=rec["name"]; cat=rec.get("category","Other")
                        ref=rec.get("ref_no","")
                        stat.markdown(f"⏳ Generating **{nm}** ... ({i+1}/{len(filtered)})")
                        png=generate_cert(nm,st.session_state.template_bytes,cn)
                        safe=nm.replace("/","_").replace("\\","_")
                        if do_png: zf.writestr(f"PNG/{cat}/{ref}_{safe}.png",png)
                        if do_pdf: zf.writestr(f"PDF/{cat}/{ref}_{safe}.pdf",cert_to_pdf(png,nm))
                        prog.progress((i+1)/len(filtered))
                stat.success(f"✅ {len(filtered)} certificates generated!")
                st.balloons()
                st.download_button("⬇️  Download All Certificates (ZIP)",bz.getvalue(),
                    file_name=f"{st.session_state.event_name.replace(' ','_')}_Certificates.zip",
                    mime="application/zip",use_container_width=True)
        st.markdown('</div>',unsafe_allow_html=True)

    # ════════════════════════════════════════════════════
    #  SOURCE B — UPLOAD NAMES FILE
    # ════════════════════════════════════════════════════
    else:
        st.markdown('<div class="card">',unsafe_allow_html=True)
        st.markdown("#### 📁 Upload Names File")

        # Format guide
        with st.expander("📋 Supported File Formats — Click to see",expanded=True):
            fc1,fc2,fc3=st.columns(3)
            with fc1:
                st.markdown("""
                <div class="card-info" style="padding:14px;text-align:center;">
                  <div style="font-size:2rem;">📗</div>
                  <b style="color:#ffd159;">Excel (.xlsx / .xls)</b><br><br>
                  <small style="color:#a8d4ff;">
                  Column A = Names<br>
                  Row 1 = Header (optional)<br><br>
                  Example:<br>
                  <code>name</code><br>
                  <code>Ali Khan</code><br>
                  <code>Fatima Zahra</code>
                  </small>
                </div>""",unsafe_allow_html=True)
            with fc2:
                st.markdown("""
                <div class="card-info" style="padding:14px;text-align:center;">
                  <div style="font-size:2rem;">📄</div>
                  <b style="color:#ffd159;">CSV (.csv)</b><br><br>
                  <small style="color:#a8d4ff;">
                  Column "name" detected<br>
                  Or use first column<br><br>
                  Example:<br>
                  <code>name,roll_no</code><br>
                  <code>Ali Khan,CS-01</code><br>
                  <code>Fatima,CS-02</code>
                  </small>
                </div>""",unsafe_allow_html=True)
            with fc3:
                st.markdown("""
                <div class="card-info" style="padding:14px;text-align:center;">
                  <div style="font-size:2rem;">📝</div>
                  <b style="color:#ffd159;">Text (.txt)</b><br><br>
                  <small style="color:#a8d4ff;">
                  One name per line<br>
                  Simplest format<br><br>
                  Example:<br>
                  <code>Ali Khan</code><br>
                  <code>Fatima Zahra</code><br>
                  <code>Ahmed Hassan</code>
                  </small>
                </div>""",unsafe_allow_html=True)

        # Sample downloads
        st.markdown("**📥 Download Sample Files:**")
        sm1,sm2,sm3=st.columns(3)
        with sm1:
            st.download_button("⬇️ Sample TXT",
                "Muhammad Ali Khan\nAyesha Siddiqui\nAhmed Hassan\nFatima Zahra\nUsman Ghani\nZara Khan".encode(),
                file_name="sample_names.txt",mime="text/plain",use_container_width=True)
        with sm2:
            st.download_button("⬇️ Sample CSV",
                "name\nMuhammad Ali Khan\nAyesha Siddiqui\nAhmed Hassan\nFatima Zahra\nUsman Ghani".encode(),
                file_name="sample_names.csv",mime="text/csv",use_container_width=True)
        with sm3:
            wb_s=openpyxl.Workbook(); ws_s=wb_s.active; ws_s.title="Names"
            ws_s["A1"]="name"
            for ni,nm in enumerate(["Muhammad Ali Khan","Ayesha Siddiqui","Ahmed Hassan","Fatima Zahra","Usman Ghani"],2):
                ws_s[f"A{ni}"]=nm
            buf_s=io.BytesIO(); wb_s.save(buf_s)
            st.download_button("⬇️ Sample Excel",buf_s.getvalue(),
                file_name="sample_names.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

        st.markdown("---")
        st.markdown("**📤 Upload Your File:**")
        names_file=st.file_uploader("",type=["xlsx","xls","csv","txt"],
                                    key="names_upl",label_visibility="collapsed")

        if names_file:
            parsed_names, err = parse_names_file(names_file)
            if err:
                st.error(f"❌ Error reading file: {err}")
                st.caption("Try saving as .xlsx or .csv and upload again.")
            elif not parsed_names:
                st.error("❌ No valid names found. Check file format above.")
            else:
                st.markdown(f'<div class="card-success">✅ <b>{len(parsed_names)} names</b> loaded from <b>{names_file.name}</b></div>',unsafe_allow_html=True)

                # Show name chips preview
                with st.expander(f"👁️ Preview names ({min(15,len(parsed_names))} shown)"):
                    chips="".join(f'<span class="chip">👤 {n}</span>' for n in parsed_names[:15])
                    if len(parsed_names)>15:
                        chips+=f'<span class="chip chip-more">+{len(parsed_names)-15} more</span>'
                    st.markdown(chips,unsafe_allow_html=True)

                # Edit names option
                with st.expander("✏️ Edit names before generating"):
                    edited_txt=st.text_area("One name per line:",
                        value="\n".join(parsed_names),height=220,key="edit_names_ta")
                    if st.button("✅ Apply Changes",key="apply_edit_btn"):
                        new_names=[l.strip() for l in edited_txt.splitlines() if l.strip()]
                        st.session_state.uploaded_names=new_names
                        st.success(f"✅ {len(new_names)} names updated!")
                        st.rerun()

                # Use edited list if applied
                final_names=(st.session_state.uploaded_names
                             if st.session_state.uploaded_names else parsed_names)

                st.markdown(f"""
                <div class="card-info" style="padding:12px;text-align:center;margin:12px 0;">
                  <b style="color:#ffd159;font-size:1.2rem;">{len(final_names)}</b>
                  <span style="color:#7ecefd;"> certificates will be generated</span>
                </div>""",unsafe_allow_html=True)

                fu1,fu2=st.columns(2)
                with fu1: do_png_u=st.checkbox("✅ PNG (Image)",value=True,key="dpu1")
                with fu2: do_pdf_u=st.checkbox("✅ PDF (Print quality)",value=True,key="dpu2")

                if st.button(f"🚀 Generate {len(final_names)} Certificates from File",
                             use_container_width=True):
                    cn=cur_cfg(); prog=st.progress(0); stat=st.empty(); bz=io.BytesIO()
                    with zipfile.ZipFile(bz,"w",zipfile.ZIP_DEFLATED) as zf:
                        for i,nm in enumerate(final_names):
                            stat.markdown(f"⏳ **{nm}** ({i+1}/{len(final_names)})")
                            png=generate_cert(nm,st.session_state.template_bytes,cn)
                            safe=nm.replace("/","_").replace("\\","_")
                            if do_png_u: zf.writestr(f"PNG/{safe}.png",png)
                            if do_pdf_u: zf.writestr(f"PDF/{safe}.pdf",cert_to_pdf(png,nm))
                            prog.progress((i+1)/len(final_names))
                    stat.success(f"✅ {len(final_names)} certificates generated!")
                    st.balloons()
                    st.session_state.uploaded_names=[]  # reset
                    st.download_button("⬇️  Download All Certificates (ZIP)",bz.getvalue(),
                        file_name="Certificates_from_File.zip",
                        mime="application/zip",use_container_width=True)
        else:
            st.markdown("""
            <div style="text-align:center;padding:30px;color:#7ecefd44;">
              <div style="font-size:3rem;margin-bottom:8px;">📂</div>
              <p>Upload a file above to get started</p>
              <small>Download sample files for reference</small>
            </div>""",unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)

# ─── TAB 6 — Backup & Security ──────────────────────────────────
with tab6:
    st.markdown("### 💾 Backup & Security")
    auth_info=load_auth()
    sc1,sc2,sc3=st.columns(3)
    sc1.metric("Algorithm","PBKDF2-SHA256","Secure")
    sc2.metric("Iterations","310,000","OWASP 2024")
    sc3.metric("Salt","256-bit","Unique")
    if "changed" in auth_info: st.success(f"✅ Password last changed: {auth_info['changed'][:10]}")
    else: st.markdown('<div class="card-warn">⚠️ Still using default password — change it from sidebar!</div>',unsafe_allow_html=True)
    st.markdown("---")
    regs_b=load_registrations()
    bc1,bc2=st.columns(2)
    with bc1:
        st.metric("Total Registrations",len(regs_b))
        st.download_button("⬇️ Download Backup ZIP",create_backup(),
            file_name=f"Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
            mime="application/zip",use_container_width=True)
        st.caption("Includes: registrations.csv + config.json")
    with bc2:
        bfiles=sorted(os.listdir(BACKUP_DIR)) if os.path.exists(BACKUP_DIR) else []
        st.markdown(f"**Server auto-backups:** `{len(bfiles)}`")
        for bf in bfiles[-5:]: st.caption(f"📁 {bf}")
    st.markdown("---")
    st.markdown("#### 🔄 Restore from CSV")
    upl_r=st.file_uploader("Upload registrations.csv:",type=["csv"])
    if upl_r:
        try:
            rdf=pd.read_csv(upl_r); st.success(f"✅ {len(rdf)} records found")
            st.dataframe(rdf.head(5),use_container_width=True)
            if st.button("⚠️ Confirm Restore (overwrites current data)"):
                rdf.to_csv(DATA_FILE,index=False); st.success("✅ Restored!"); st.rerun()
        except Exception as e: st.error(f"❌ {e}")
    st.markdown("---")
    st.markdown("#### ⚠️ Danger Zone")
    with st.expander("🗑️ Delete All Registrations (irreversible)"):
        conf=st.text_input("Type DELETE to confirm:")
        if st.button("🗑️ Delete All") and conf=="DELETE":
            bak=create_backup(); clear_registrations()
            st.warning("✅ Deleted! Auto-backup taken.")
            st.download_button("⬇️ Pre-delete Backup",bak,
                file_name="pre_delete.zip",mime="application/zip")
            st.rerun()

# ─── TAB 7 — Deploy Guide ───────────────────────────────────────
with tab7:
    st.markdown("""
<div class="card">

## ☁️ Deploy Guide — GitHub + Streamlit Cloud (Free)

### Files to push:
```
app.py
requirements.txt
```

### Git commands:
```bash
cd d:/Avalon.AI
git add app.py requirements.txt
git commit -m "v7 - file upload + premium UI"
git push
```

### Deploy steps:
1. [share.streamlit.io](https://share.streamlit.io) → GitHub login
2. New App → repo → `app.py` → Deploy
3. Copy URL → Admin Sidebar → Save Settings → Generate QR ✅

### Default password: `Admin@2025` — change immediately after login!

### 🔒 To save password permanently (Streamlit Cloud):
After changing password, copy the TOML code shown in sidebar  
→ Streamlit Dashboard → App → ⚙️ Settings → Secrets → Paste → Save

</div>
""",unsafe_allow_html=True)

# ─── TAB 8 — Developer ──────────────────────────────────────────
with tab8:
    st.markdown("""
<div class="dev-card">
  <div style="font-size:5.5rem;margin-bottom:4px;">👨‍💻</div>
  <div class="dev-name">Abdul Samad</div>
  <div class="dev-title">Web Developer  •  AI/ML Enthusiast  •  Entreprenuer Enthusiast </div>
  <hr class="dev-hr">
  <p style="color:#c0d4ee;font-size:1rem;margin:4px 0;">🎓 <b>BS Computer Science</b></p>
  <p style="color:#c0d4ee;font-size:1rem;margin:4px 0;">Shaheed Benazir Bhutto University (SBBU), Nawabshah, Sindh, Pakistan</p>
  <hr class="dev-hr">
  <div style="color:#7ecefd;font-weight:700;margin-bottom:8px;font-size:.95rem;">🔗 Connect</div>
  <div class="soc-row">
    <a class="soc-btn" href="https://instagram.com/isamad?ighs=MThwaXU3N2QwdGplcg==" target="_blank"
       style="background:linear-gradient(45deg,#f09433,#e6683c,#dc2743,#cc2366,#bc1888);color:white;">📷 Instagram</a>
    <a class="soc-btn" href="https://facebook.com/isamadrind" target="_blank"
       style="background:#1877F2;color:white;">📘 Facebook</a>
    <a class="soc-btn" href="https://linkedin.com/in/abdul-samad-rind-842724338?utm_source=share&utm_campaign=share_via&utm_content=profile&utm_medium=android_app" target="_blank"
       style="background:#0A66C2;color:white;">💼 LinkedIn</a>
    <a class="soc-btn" href="https://tiktok.com/@isamadrind" target="_blank"
       style="background:linear-gradient(90deg,#010101,#69C9D0);color:white;">🎵 TikTok</a>
    <a class="soc-btn" href="tel:+92313-0328282"
       style="background:#25D366;color:white;">📞 +92-313-0328282</a>
  </div>
  <hr class="dev-hr">
  <div style="color:#7ecefd;font-weight:700;margin-bottom:10px;font-size:.95rem;">💡 Skills</div>
  <div class="skill-row">
    <span class="skill">UI/UX Design</span><span class="skill">Web Developer</span>
    <span class="skill">Python</span><span class="skill">Streamlit</span>
    <span class="skill">Deep Learning</span><span class="skill">Machine Learning</span>
    <span class="skill">Web Development</span><span class="skill">AI</span>
    <span class="skill">FastAPI</span><span class="skill">SQL & Databases</span>
    <span class="skill">Pandas & NumPy</span><span class="skill">Git & GitHub</span>
    <span class="skill">UI/UX Design</span><span class="skill">Pillow / PIL</span>
    <span class="skill">OpenCV</span>
  </div>
  <hr class="dev-hr">
  <div style="color:#7ecefd;font-size:.92rem;line-height:1.9;">
    <b style="color:#ffd159;">QR Certificate System v7.0</b><br>
    Built with Python • Streamlit • Pillow • ReportLab • OpenPyXL
  </div>
  <p style="color:#7ecefd33;font-size:.82rem;margin-top:20px;">
    © 2026 Abdul Samad — All Rights Reserved<br>
    Developed with ❤️ at SBBU Nawabshah
  </p>
</div>
""",unsafe_allow_html=True)
    st.markdown("### 📬 Contact")
    gc1,gc2,gc3=st.columns(3)
    with gc1: st.markdown('<div class="card" style="text-align:center;"><div style="font-size:2rem;">📧</div><p style="color:#ffd159;font-weight:700;">Email</p><p style="color:#7ecefd;">asamad009@outlook.com</p></div>',unsafe_allow_html=True)
    with gc2: st.markdown('<div class="card" style="text-align:center;"><div style="font-size:2rem;">🌐</div><p style="color:#ffd159;font-weight:700;">Portfolio</p><p style="color:#7ecefd;">isamadrind.kesug.com</p></div>',unsafe_allow_html=True)
    with gc3: st.markdown('<div class="card" style="text-align:center;"><div style="font-size:2rem;">📍</div><p style="color:#ffd159;font-weight:700;">Location</p><p style="color:#7ecefd;">Nawabshah, Sindh, Pakistan</p></div>',unsafe_allow_html=True)

# ─── TAB 9 — README ─────────────────────────────────────────────
with tab9:
    st.markdown("""
<div class="card">

# 📖 QR Certificate Generator Pro v3.01 — User Guide

## 🎯 What Is This?
A complete event management platform — QR registration → instant invitation card → bulk certificates.

---

## 🚀 Certificate Generation — 2 Sources

### Source 1: Online Registrations
- Attendees scan QR → fill form → auto-saved to database
- Go to **🚀 Bulk Generate** → select **Online Registrations**
- Filter by category → Generate → Download ZIP

### Source 2: Upload Names File *(NEW in v7)*
- Have an existing list? Upload it directly!
- **Supported formats:**
  - `.xlsx` / `.xls` — Excel file (Column A = names)
  - `.csv` — CSV file (column named "name" or first column)
  - `.txt` — Plain text (one name per line)
- Preview names before generating
- Edit/add/remove names manually
- Download sample files for reference

---

## 🎫 Invitation Card
- **8 themes:** Royal Gold, Midnight Blue, Crimson Elite, Emerald, Obsidian Gold, Ocean Sapphire, Violet Luxury, Rose Gold
- **Short Reg No:** P-0001, TC-0005, SP-0012
- **Smart phrases** by category (Teacher/Speaker/Business/Student)
- **1-3 logos** supported
- **1080×1620px PNG** — perfect for mobile & social media

## 🔒 Security
- PBKDF2-SHA256, 310,000 iterations, 256-bit salt
- Password stored hashed in `auth.json` — never plaintext
- Default: `Admin@2025` — change immediately!

## 💾 Data
- `registrations.csv` — all data, survives restarts
- `config.json` — settings
- Daily auto-backup to `backups/` folder
- Manual backup/restore in Tab 6

## ⚙️ Requirements
```
streamlit pillow qrcode[pil] reportlab openpyxl pandas
```

*© 2026 Abdul Samad — SBBU Nawabshah*

</div>
""",unsafe_allow_html=True)

# ── Footer ───────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<p style="text-align:center;color:#7ecefd22;font-size:.82rem;">© Certificate Generoator Pro v3.01 | Developed by Abdul Samad | SBBU Nawabshah, Pakistan</p>',unsafe_allow_html=True)
